from django.shortcuts import render, redirect
from .forms import AdmissionEnquiryForm
from .models import Administrator
import openpyxl
import os

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def enquiry_form(request):
    if request.method == "POST":
        form = AdmissionEnquiryForm(request.POST)

        if form.is_valid():
            print("FORM VALID ✅")

                        # Get path entered in form
            new_path = form.cleaned_data.get('excel_file_path')

            # If user entered a new path → save it
            if new_path and new_path.strip() != "":
                file_path = new_path.strip()
                request.session['excel_file_path'] = file_path
            else:
                # Otherwise use previous saved path
                file_path = request.session.get('excel_file_path')
            if not file_path:
                print("Excel path not provided ❌")
                return redirect('/')

            enquiry = form.save()

            try:
                # Create folder if not exists
                folder = os.path.dirname(file_path)
                if folder and not os.path.exists(folder):
                    os.makedirs(folder)

                headers = [
                    "Student Name",
                    "Contact",
                    "Email",
                    "Mode Of Enquiry",
                    "Course Name",
                    "Enquiry Date",
                    "Address",
                    "Enquiry Handled By",
                    "Remarks",
                    "Reference",
                    "Reason Not To Join",
                    "Educational"
                ]

                # Create workbook if not exists
                if not os.path.exists(file_path):
                    wb = openpyxl.Workbook()
                    sheet = wb.active
                    sheet.title = "Admission Enquiries"
                    sheet.append(headers)
                    wb.save(file_path)

                wb = openpyxl.load_workbook(file_path)
                sheet = wb.active

                # Ensure header exists
                if sheet.max_row == 0 or sheet["A1"].value != "Student Name":
                    sheet.delete_rows(1, sheet.max_row)
                    sheet.append(headers)

                # Append enquiry
                sheet.append([
                    enquiry.student_name,
                    enquiry.contact,
                    enquiry.email,
                    enquiry.mode_of_enquiry,
                    enquiry.course_name,
                    str(enquiry.enquiry_date),
                    enquiry.address,
                    enquiry.enquiry_handled_by,
                    enquiry.remarks,
                    enquiry.reference,
                    enquiry.reason_not_to_join,
                    enquiry.educational
                ])

                # Header styling
                header_fill = PatternFill(
                    start_color="4F81BD",
                    end_color="4F81BD",
                    fill_type="solid"
                )

                for col in range(1, len(headers) + 1):
                    cell = sheet.cell(row=1, column=col)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = header_fill

                # Borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for row in sheet.iter_rows():
                    for cell in row:
                        cell.border = thin_border

                # Auto column width
                for column_cells in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column_cells[0].column)

                    for cell in column_cells:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))

                    sheet.column_dimensions[column_letter].width = max_length + 3

                # Create table
                last_row = sheet.max_row
                last_col = sheet.max_column
                table_range = f"A1:{get_column_letter(last_col)}{last_row}"

                if sheet._tables:
                    sheet._tables.clear()

                table = Table(displayName="AdmissionTable", ref=table_range)

                style = TableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )

                table.tableStyleInfo = style
                sheet.add_table(table)

                sheet.freeze_panes = "A2"

                wb.save(file_path)

                print(f"Data saved successfully to {file_path} ✅")

            except Exception as e:
                print("Excel Error ❌:", e)

            return redirect('/')

        else:
            print("FORM ERRORS ❌:", form.errors)

    else:
        form = AdmissionEnquiryForm()

    return render(request, 'form.html', {
    'form': form,
    'saved_path': request.session.get('excel_file_path')
})


def enquiry_list(request):
    enquiries = Administrator.objects.all().order_by('-id')
    return render(request, 'list.html', {'enquiries': enquiries})


def delete_enquiry(request, id):
    enquiry = Administrator.objects.get(id=id)
    enquiry.delete()
    return redirect('/list/')